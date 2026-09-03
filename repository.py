import pandas as pd
import uuid
import os
import shutil
import sqlite3
import datetime
import time
import re
from utils import debug_error

# ---------------------------------------------------------------------------
# Column name constants used across repository and UI layers.
# Import these in other modules rather than redefining them.
# ---------------------------------------------------------------------------
REVIEWED_COLUMN = "Reviewed"
REVIEWED_AT_COLUMN = "ReviewedAt"
ONLINE_EXISTS_COLUMN = "Online_Images_Exist"


_TRUTHY_SET = frozenset({"true", "1", "yes", "y", "t", "on", 1, 1.0, True})
_FALSY_SET = frozenset({"false", "0", "no", "n", "f", "off", "", 0, 0.0, False})
_SCIENTIFIC_REGEX = re.compile(r'^[+-]?(?:\d+\.?\d*|\.\d+)[eE][+-]?\d+$')
_FORMULA_ERRORS = frozenset({
    '#REF!', '#VALUE!', '#N/A', '#DIV/0!', '#NAME?', '#NUM!', '#NULL!', '#CALC!', '#SPILL!'
})


def _coerce_bool_series(series: pd.Series, default: bool = False) -> pd.Series:
    """Safely coerce a pandas Series to boolean without Python string truthiness bugs."""
    if series is None or series.empty:
        return series

    if pd.api.types.is_bool_dtype(series):
        return series.fillna(default)
    if pd.api.types.is_numeric_dtype(series):
        return series.fillna(default).astype(bool)

    s = series.astype(str).str.strip().str.lower()

    # Vectorized true/false matching
    is_true = s.isin(_TRUTHY_SET)
    is_false = s.isin(_FALSY_SET)

    # Elements not found in either set fallback to default if they were originally NaN, otherwise standard truthiness.
    # To handle the truthiness correctly for the fallback case:
    # We will construct the boolean series using numpy/pandas operations.

    # Create the result series, initializing with the default value
    res = pd.Series(default, index=series.index, dtype=bool)

    res.loc[is_true] = True
    res.loc[is_false] = False

    # For anything else (neither in truthy nor falsy set):
    # If the original was not NaN/None, apply standard boolean casting.
    mask_other = ~is_true & ~is_false & ~series.isna() & (series.astype(str) != 'None')
    if mask_other.any():
        res.loc[mask_other] = series.loc[mask_other].astype(bool)

    return res


def _normalize_single_id(val: str) -> str:
    """Normalize a single ObjectID string: strips float .0, converts scientific notation, strips formula errors."""
    if not val or val in _FORMULA_ERRORS:
        return ""
    # Strip trailing .0 if integer representation
    if val.endswith(".0") and (val[:-2].isdigit() or (val.startswith("-") and val[1:-2].isdigit())):
        return val[:-2]
    # Check scientific notation (e.g. 1E+05, 1.0e+05)
    if _SCIENTIFIC_REGEX.match(val):
        try:
            f = float(val)
            if f.is_integer():
                return str(int(f))
            return f"{f:f}".rstrip('0').rstrip('.')
        except (ValueError, OverflowError):
            pass
    return val


def _normalize_object_id_series(series: pd.Series) -> pd.Series:
    """Normalize ObjectID series preserving leading zeros, expanding scientific notation, and stripping float .0 noise."""
    if series is None or series.empty:
        return series

    # Convert to string and handle standard missing values.
    # fillna("") first so np.nan doesn't become the string "nan".
    s = series.fillna("").astype(str).str.strip()
    s = s.replace(['nan', 'None', '<NA>'], '')

    return s.map(_normalize_single_id)


def _deduplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure all column names in a DataFrame are unique, stripped strings."""
    if df is None or df.empty:
        return df
    cols = []
    seen = {}
    for col in df.columns:
        c_str = str(col).strip()
        if c_str in seen:
            seen[c_str] += 1
            cols.append(f"{c_str}.{seen[c_str]}")
        else:
            seen[c_str] = 0
            cols.append(c_str)
    if cols != list(df.columns):
        df = df.copy(deep=False)
        df.columns = cols
    return df


def _detect_and_promote_header(df: pd.DataFrame, expected_col: str = "ObjectID", max_scan_rows: int = 15) -> pd.DataFrame:
    """Detect if the true header row is offset downward due to empty or title rows."""
    if df is None or df.empty:
        return df

    # Case-insensitive check if expected_col is already in columns
    col_map = {str(c).strip().lower(): c for c in df.columns}
    target_lower = expected_col.strip().lower()
    if target_lower in col_map:
        if col_map[target_lower] != expected_col:
            df = df.rename(columns={col_map[target_lower]: expected_col})
        return df

    # Scan the first max_scan_rows to find where expected_col appears
    for row_idx in range(min(len(df), max_scan_rows)):
        row_vals = [str(v).strip().lower() for v in df.iloc[row_idx].values if pd.notna(v)]
        if target_lower in row_vals:
            new_header = [str(c).strip() if pd.notna(c) else f"Unnamed_{i}" for i, c in enumerate(df.iloc[row_idx])]
            df_promoted = df.iloc[row_idx + 1:].copy()
            df_promoted.columns = new_header
            df_promoted.reset_index(drop=True, inplace=True)
            col_map_new = {str(c).strip().lower(): c for c in df_promoted.columns}
            if target_lower in col_map_new and col_map_new[target_lower] != expected_col:
                df_promoted = df_promoted.rename(columns={col_map_new[target_lower]: expected_col})
            return df_promoted

    return df


def _clean_trailing_and_blank_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Drop unnamed columns that are entirely blank and drop ghost rows that are completely empty."""
    if df is None or df.empty:
        return df

    # 1. Drop unnamed/blank columns that contain no data
    cols_to_drop = []
    for col in df.columns:
        col_str = str(col).strip()
        if col_str.startswith("Unnamed:") or col_str.startswith("Unnamed_") or col_str == "":
            s = df[col]
            if s.isna().all() or (s.fillna("").astype(str).str.strip() == "").all():
                cols_to_drop.append(col)
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)

    # 2. Drop completely empty ghost rows (all NaN or all empty string)
    if not df.empty:
        is_empty_row = df.isna() | (df.astype(str).apply(lambda col: col.str.strip() == ""))
        df = df.loc[~is_empty_row.all(axis=1)].copy()
        df.reset_index(drop=True, inplace=True)

    return df


def _open_excel_reader(path: str) -> pd.ExcelFile:
    """Open Excel workbook with calamine for high performance, falling back to openpyxl."""
    try:
        return pd.ExcelFile(path, engine="calamine")
    except Exception:
        return pd.ExcelFile(path, engine="openpyxl")


def _find_sheet_name(sheet_names: list[str], target_name: str) -> str | None:
    """Find sheet name with case and whitespace insensitivity."""
    if not target_name or not sheet_names:
        return None
    if target_name in sheet_names:
        return target_name

    target_clean = target_name.strip().lower()
    mapping = {s.strip().lower(): s for s in sheet_names}
    return mapping.get(target_clean)


def _normalise_dataframes(df_reg, df_obs, config):
    """Apply consistent type coercion and column defaults to the loaded dataframes.

    Both ExcelRepository and SQLiteRepository call this after reading raw data so
    that the rest of the application always receives dataframes in a known shape.

    Operations performed:
    - Strips and normalizes ObjectID to str in df_reg, df_obs.
    - Ensures all problem columns exist in df_obs and are bool.
    - Ensures all location columns exist in df_obs and are str.
    - Ensures Images_Missing, Images_Problem, Reviewed, ReviewedAt,
      Online_Images_Exist columns exist with appropriate defaults.
    - Fills NaN in all non-ObjectID df_reg columns with empty string.
    - Ensures UID and ProblemDescription columns exist in df_reg.
    - Auto-generates short UIDs for any row that is missing one.

    Args:
        df_reg: Registration dataframe.
        df_obs: Observation dataframe.
        config: Database config dict containing ui_sections.

    Returns:
        (df_reg, df_obs) — the normalised DataFrames.
    """
    sections = config.get("ui_sections", {})
    problem_columns = [f["name"] for f in sections.get("problems", [])]
    location_columns = [f["name"] for f in sections.get("location", [])]
    registration_columns = [f["name"] for f in sections.get("registration", [])]

    # --- ObjectID ---
    if "ObjectID" in df_reg.columns:
        df_reg["ObjectID"] = _normalize_object_id_series(df_reg["ObjectID"])
    if "ObjectID" in df_obs.columns:
        df_obs["ObjectID"] = _normalize_object_id_series(df_obs["ObjectID"])

    # Ensure df_obs has ObjectID column if df_reg does
    if "ObjectID" in df_reg.columns and "ObjectID" not in df_obs.columns:
        df_obs["ObjectID"] = pd.Series(dtype=str)

    # Ensure all ObjectIDs in df_reg exist in df_obs
    if "ObjectID" in df_reg.columns and "ObjectID" in df_obs.columns:
        reg_ids = df_reg["ObjectID"].dropna().unique()
        obs_ids = set(df_obs["ObjectID"].dropna().unique())
        missing_obs_ids = [rid for rid in reg_ids if rid not in obs_ids]
        if missing_obs_ids:
            missing_df = pd.DataFrame({"ObjectID": missing_obs_ids})
            df_obs = pd.concat([df_obs, missing_df], ignore_index=True)

    # --- Registration: batch-insert missing registration columns (anti-fragmentation) ---
    new_reg = {col: "" for col in registration_columns if col not in df_reg.columns}
    if "UID" not in df_reg.columns and "UID" not in new_reg:
        new_reg["UID"] = ""
    if "ProblemDescription" not in df_reg.columns and "ProblemDescription" not in new_reg:
        new_reg["ProblemDescription"] = ""

    if new_reg:
        new_reg_df = pd.DataFrame(new_reg, index=df_reg.index)
        df_reg = pd.concat([df_reg, new_reg_df], axis=1)

    # --- Observation: batch-insert missing observation columns (anti-fragmentation) ---
    new_obs = {}
    for col in problem_columns:
        if col not in df_obs.columns:
            new_obs[col] = False
    for col in location_columns:
        if col not in df_obs.columns:
            new_obs[col] = ""

    if "Images_Missing" not in df_obs.columns:
        new_obs["Images_Missing"] = True
    if "Images_Problem" not in df_obs.columns:
        new_obs["Images_Problem"] = False
    if REVIEWED_COLUMN not in df_obs.columns:
        new_obs[REVIEWED_COLUMN] = False
    if REVIEWED_AT_COLUMN not in df_obs.columns:
        new_obs[REVIEWED_AT_COLUMN] = ""
    if ONLINE_EXISTS_COLUMN not in df_obs.columns:
        new_obs[ONLINE_EXISTS_COLUMN] = False

    if new_obs:
        new_obs_df = pd.DataFrame(new_obs, index=df_obs.index)
        df_obs = pd.concat([df_obs, new_obs_df], axis=1)

    for pcol in problem_columns:
        if pcol in df_obs.columns:
            df_obs[pcol] = _coerce_bool_series(df_obs[pcol], default=False)

    if location_columns:
        df_obs[location_columns] = df_obs[location_columns].fillna("").astype(object)

    for field in sections.get("location", []):
        if field.get("type") == "checkbox":
            col = field["name"]
            if col in df_obs.columns:
                s = df_obs[col]
                is_true = s.isin([True, "True", 1, "1", "yes", "true", "YES", "TRUE"])
                df_obs[col] = is_true.map({True: "True", False: "False"}).astype(str)

    df_obs["Images_Missing"] = _coerce_bool_series(df_obs["Images_Missing"], default=True)
    df_obs["Images_Problem"] = _coerce_bool_series(df_obs["Images_Problem"], default=False)
    df_obs[REVIEWED_COLUMN] = _coerce_bool_series(df_obs[REVIEWED_COLUMN], default=False)
    if ONLINE_EXISTS_COLUMN in df_obs.columns:
        df_obs[ONLINE_EXISTS_COLUMN] = _coerce_bool_series(df_obs[ONLINE_EXISTS_COLUMN], default=False)
    df_obs[REVIEWED_AT_COLUMN] = df_obs[REVIEWED_AT_COLUMN].fillna("").astype(object)

    # --- Registration: fill NaN, ensure types and UIDs ---
    cols_to_fill = [col for col in df_reg.columns if col != "ObjectID"]
    if cols_to_fill:
        df_reg[cols_to_fill] = df_reg[cols_to_fill].fillna("").astype(object)

    df_reg["ProblemDescription"] = df_reg["ProblemDescription"].astype(object)

    # Generate short UIDs for any row that is missing one
    missing_uid = df_reg["UID"].isna() | (df_reg["UID"].astype(str).str.strip() == "")
    missing_count = missing_uid.sum()
    if missing_count > 0:
        df_reg.loc[missing_uid, "UID"] = [
            uuid.uuid4().hex[:8] for _ in range(missing_count)
        ]

    # Auto-repair legacy datasets where Other_problem was corrupted (set to True on virtually all rows)
    if "Other_problem" in df_obs.columns and not df_obs.empty and len(df_obs) > 1:
        other_true_series = _coerce_bool_series(df_obs["Other_problem"], default=False)
        if other_true_series.mean() >= 0.9:  # Corrupted dataset threshold (e.g. >= 90% rows True)
            has_desc = pd.Series(False, index=df_reg.index)
            for col in ["ProblemDescription", "Comment", "Observation"]:
                if col in df_reg.columns:
                    has_desc |= (df_reg[col].fillna("").astype(str).str.strip() != "")
            if "ObjectID" in df_reg.columns and "ObjectID" in df_obs.columns:
                desc_ids = set(df_reg.loc[has_desc, "ObjectID"].dropna().unique())
                df_obs["Other_problem"] = other_true_series & df_obs["ObjectID"].isin(desc_ids)
            else:
                df_obs["Other_problem"] = other_true_series & has_desc

    return df_reg, df_obs

def _normalise_unvalidated_dataframe(df_unval):
    """Ensure the unvalidated source dataframe has the required columns: ObjectID, Field_Name, Unvalidated_Comment."""
    required_cols = ["ObjectID", "Field_Name", "Unvalidated_Comment"]
    if df_unval is None or (isinstance(df_unval, pd.DataFrame) and df_unval.empty):
        return pd.DataFrame(columns=required_cols)

    df = df_unval.copy()
    if "ObjectID" not in df.columns and getattr(df.index, "name", None) == "ObjectID":
        df = df.reset_index()

    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        for c in missing:
            df[c] = ""

    df = df[required_cols]
    df["ObjectID"] = _normalize_object_id_series(df["ObjectID"])
    df["Field_Name"] = df["Field_Name"].fillna("").astype(str).str.strip()
    df["Unvalidated_Comment"] = df["Unvalidated_Comment"].fillna("").astype(str).str.strip()
    # Filter out completely blank rows
    mask_nonblank = (df["ObjectID"] != "") | (df["Field_Name"] != "") | (df["Unvalidated_Comment"] != "")
    df = df[mask_nonblank]
    return df


def _normalise_log_dataframe(df_log):
    """Ensure the log dataframe has all required columns.

    This adds backwards compatibility for older databases that may not have
    the new section-specific logging columns, strips any internal session
    tracking columns, and guarantees canonical column ordering.
    """
    required_cols = [
        "Timestamp", "Action", "Reviewed", "ObjectID",
        "ChangedFields", "ChangedValues",
        "ProblemsChanged", "ProblemsChangedValues",
        "LocationChanged", "LocationChangedValues",
        "User", "SourceFile", "OutputFile"
    ]
    
    if df_log is None or df_log.empty:
        return pd.DataFrame(columns=required_cols)
        
    df_clean = df_log.copy()
    missing_cols = [col for col in required_cols if col not in df_clean.columns]
    if missing_cols:
        new_cols_df = pd.DataFrame({col: "" for col in missing_cols}, index=df_clean.index)
        df_clean = pd.concat([df_clean, new_cols_df], axis=1)
            
    return df_clean[required_cols].copy()


class ExcelRepository:
    @staticmethod
    def load_excel(path, config):
        """Load a database from an Excel (.xlsx) file.

        Reads the Registration, Observation, Photo, Log, and Unvalidated_source sheets.
        The Photo, Log, and Unvalidated_source sheets are optional — empty dataframes with the
        correct columns are returned if the sheets are missing.
        All dataframes are normalised via _normalise_dataframes before being
        returned, so callers always receive data in a consistent shape.

        Args:
            path:   Absolute path to the .xlsx file.
            config: Database config dict (from config.DATABASE_CONFIGS).

        Returns:
            (df_reg, df_obs, df_photo, df_log, df_unvalidated) — five pandas DataFrames.
        """
        sheets = config["sheets"]
        sections = config["ui_sections"]

        mapped_fields = [
            f["maps_to"]
            for f in sections["problems"]
            if "maps_to" in f
        ]

        with _open_excel_reader(path) as xls:
            sheet_names = xls.sheet_names

            # Read Registration sheet
            target_reg = sheets.get("reg", "Registration")
            matched_reg = _find_sheet_name(sheet_names, target_reg)
            if matched_reg:
                df_reg = pd.read_excel(xls, sheet_name=matched_reg)
            else:
                df_reg = pd.read_excel(xls, sheet_name=target_reg)  # Let it raise with original target name
            df_reg = _clean_trailing_and_blank_columns(_deduplicate_columns(_detect_and_promote_header(df_reg, "ObjectID")))

            # Read Observation sheet
            target_obs = sheets.get("obs", "Observation")
            matched_obs = _find_sheet_name(sheet_names, target_obs)
            if matched_obs:
                df_obs = pd.read_excel(xls, sheet_name=matched_obs)
            else:
                df_obs = pd.DataFrame(columns=["ObjectID"])
            df_obs = _clean_trailing_and_blank_columns(_deduplicate_columns(_detect_and_promote_header(df_obs, "ObjectID")))

            # Read Photo sheet (optional)
            target_photo = sheets.get("photo", "Photo")
            matched_photo = _find_sheet_name(sheet_names, target_photo)
            if matched_photo:
                df_photo = pd.read_excel(xls, sheet_name=matched_photo)
            else:
                df_photo = pd.DataFrame(columns=["ObjectID"])
            if not df_photo.empty:
                df_photo = _clean_trailing_and_blank_columns(_deduplicate_columns(_detect_and_promote_header(df_photo, "ObjectID")))

            # Ensure mapped registration fields exist before normalisation
            missing_mapped_fields = [col for col in mapped_fields if col not in df_reg.columns]
            if missing_mapped_fields:
                new_mapped_df = pd.DataFrame({col: "" for col in missing_mapped_fields}, index=df_reg.index)
                df_reg = pd.concat([df_reg, new_mapped_df], axis=1)

            # Normalise both main dataframes
            df_reg, df_obs = _normalise_dataframes(df_reg, df_obs, config)

            if not df_photo.empty and "ObjectID" in df_photo.columns:
                df_photo["ObjectID"] = _normalize_object_id_series(df_photo["ObjectID"])

            # Read Log sheet (optional)
            target_log = sheets.get("log", "Log")
            matched_log = _find_sheet_name(sheet_names, target_log)
            if matched_log:
                df_log = pd.read_excel(xls, sheet_name=matched_log)
            else:
                df_log = pd.DataFrame()
            if not df_log.empty:
                df_log = _clean_trailing_and_blank_columns(_deduplicate_columns(df_log))

            # Read Unvalidated_source sheet (optional)
            target_unval = sheets.get("unvalidated", "Unvalidated_source")
            matched_unval = _find_sheet_name(sheet_names, target_unval)
            if matched_unval:
                df_unvalidated = pd.read_excel(xls, sheet_name=matched_unval)
            else:
                df_unvalidated = pd.DataFrame(columns=["ObjectID", "Field_Name", "Unvalidated_Comment"])
            if not df_unvalidated.empty:
                df_unvalidated = _clean_trailing_and_blank_columns(_deduplicate_columns(df_unvalidated))
            df_unvalidated = _normalise_unvalidated_dataframe(df_unvalidated)
            
        df_log = _normalise_log_dataframe(df_log)

        return df_reg, df_obs, df_photo, df_log, df_unvalidated

    @staticmethod
    def save_excel(path, config, df_reg=None, df_obs=None, df_log=None, df_photo=None, df_unvalidated=None, progress_callback=None):
        return SQLiteRepository.export_to_excel(None, path, config, progress_callback=progress_callback,
                                                df_reg=df_reg, df_obs=df_obs, df_log=df_log, df_photo=df_photo,
                                                df_unvalidated=df_unvalidated)


class SQLiteRepository:
    @staticmethod
    def load_sqlite(path, config):
        """Load a database from a SQLite (.db) file.

        Reads the Registration, Observation, Photo, Log, and Unvalidated_source tables.
        All tables are optional — empty dataframes with correct columns are
        returned for any missing table.
        All dataframes are normalised via _normalise_dataframes before being
        returned, so callers always receive data in a consistent shape.

        Args:
            path:   Absolute path to the .db file.
            config: Database config dict (from config.DATABASE_CONFIGS).

        Returns:
            (df_reg, df_obs, df_photo, df_log, df_unvalidated) — five pandas DataFrames.
        """
        conn = sqlite3.connect(path, timeout=30.0)
        try:
            try:
                conn.execute("PRAGMA journal_mode = WAL;")
            except sqlite3.OperationalError:
                pass
            conn.execute("PRAGMA busy_timeout = 30000;")

            try:
                df_reg = pd.read_sql("SELECT * FROM Registration", conn)
            except Exception as e:
                debug_error("load_sqlite: Registration table missing or unreadable", str(e))
                df_reg = pd.DataFrame()

            try:
                df_obs = pd.read_sql("SELECT * FROM Observation", conn)
            except Exception as e:
                debug_error("load_sqlite: Observation table missing or unreadable", str(e))
                df_obs = pd.DataFrame()

            # Check optional tables using sqlite_master
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Photo'")
            photo_exists = cursor.fetchone() is not None

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Log'")
            log_exists = cursor.fetchone() is not None

            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND (name='Unvalidated_source' OR name='unvalidated_source')")
            unval_row = cursor.fetchone()

            if photo_exists:
                try:
                    df_photo = pd.read_sql("SELECT * FROM Photo", conn)
                except Exception as e:
                    debug_error("load_sqlite: Photo table unreadable", str(e))
                    df_photo = pd.DataFrame(columns=["ObjectID"])
            else:
                df_photo = pd.DataFrame(columns=["ObjectID"])

            if log_exists:
                try:
                    df_log = pd.read_sql("SELECT * FROM Log", conn)
                except Exception as e:
                    debug_error("load_sqlite: Log table unreadable", str(e))
                    df_log = pd.DataFrame()
            else:
                df_log = pd.DataFrame()

            if unval_row:
                try:
                    df_unvalidated = pd.read_sql(f'SELECT * FROM "{unval_row[0]}"', conn)
                except Exception as e:
                    debug_error("load_sqlite: Unvalidated_source table unreadable", str(e))
                    df_unvalidated = pd.DataFrame(columns=["ObjectID", "Field_Name", "Unvalidated_Comment"])
            else:
                df_unvalidated = pd.DataFrame(columns=["ObjectID", "Field_Name", "Unvalidated_Comment"])
                
            df_log = _normalise_log_dataframe(df_log)
            df_unvalidated = _normalise_unvalidated_dataframe(df_unvalidated)
        finally:
            conn.close()

        # Normalise both main dataframes
        df_reg, df_obs = _normalise_dataframes(df_reg, df_obs, config)

        if "ObjectID" in df_photo.columns:
            df_photo["ObjectID"] = df_photo["ObjectID"].astype(str).str.strip()

        return df_reg, df_obs, df_photo, df_log, df_unvalidated

    @staticmethod
    def save_sqlite(path, df_reg, df_obs, df_photo, df_log, df_unvalidated=None):
        """Write all dataframes to a SQLite database file.

        Replaces the Registration, Observation, Photo, Log, and Unvalidated_source tables entirely.
        The Photo, Log, and Unvalidated_source tables are only written if they are non-empty.

        Args:
            path:           Absolute path to the .db file (created if it does not exist).
            df_reg:         Registration dataframe.
            df_obs:         Observation dataframe.
            df_photo:       Photo dataframe.
            df_log:         Log dataframe.
            df_unvalidated: Unvalidated source dataframe (optional).
        """
        conn = sqlite3.connect(path, timeout=30.0)
        try:
            # Optimize transaction overhead and SQLite performance pragmas
            conn.execute("PRAGMA journal_mode = WAL;")
            conn.execute("PRAGMA synchronous = NORMAL;")
            conn.execute("PRAGMA busy_timeout = 30000;")

            df_reg_save = _deduplicate_columns(df_reg.copy())
            if "ObjectID" not in df_reg_save.columns:
                df_reg_save = df_reg_save.reset_index()

            df_obs_save = _deduplicate_columns(df_obs.copy())
            if "ObjectID" not in df_obs_save.columns:
                df_obs_save = df_obs_save.reset_index()

            df_photo_save = _deduplicate_columns(df_photo.copy()) if df_photo is not None else pd.DataFrame()
            if not df_photo_save.empty and "ObjectID" not in df_photo_save.columns:
                df_photo_save = df_photo_save.reset_index()

            with conn:
                cursor = conn.cursor()
                cursor.execute("SELECT m.name, p.name FROM sqlite_master m LEFT JOIN pragma_table_info(m.name) p WHERE m.type='table';")

                existing_tables = set()
                db_table_cols = {}
                for t_name, c_name in cursor.fetchall():
                    existing_tables.add(t_name)
                    if c_name is not None:
                        db_table_cols.setdefault(t_name, set()).add(c_name)

                for table_name, df_save in [("Registration", df_reg_save), ("Observation", df_obs_save)]:
                    if table_name in existing_tables:
                        # Table exists, check if columns match (so we don't break on schema changes)
                        cursor = conn.cursor()

                        if not table_name.isidentifier():
                            raise ValueError(f"Invalid table name: {table_name}")

                        db_cols = db_table_cols.get(table_name, set())
                        df_cols = set(df_save.columns)
                        if df_cols.issubset(db_cols):
                            cursor.execute(f'DELETE FROM "{table_name}";')  # nosec B608
                            df_save.to_sql(table_name, conn, if_exists="append", index=False)
                        else:
                            # Fallback if columns differ (e.g. schema migration)
                            df_save.to_sql(table_name, conn, if_exists="replace", index=False)
                    else:
                        df_save.to_sql(table_name, conn, if_exists="replace", index=False)

                if not df_photo_save.empty:
                    if "Photo" in existing_tables:
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM Photo;")
                        df_photo_save.to_sql("Photo", conn, if_exists="append", index=False)
                    else:
                        df_photo_save.to_sql("Photo", conn, if_exists="replace", index=False)

                if df_log is not None and not df_log.empty:
                    df_log_save = _normalise_log_dataframe(df_log)
                    if "Log" in existing_tables:
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM Log;")
                        df_log_save.to_sql("Log", conn, if_exists="append", index=False)
                    else:
                        df_log_save.to_sql("Log", conn, if_exists="replace", index=False)

                if df_unvalidated is not None and not df_unvalidated.empty:
                    df_unval_save = _deduplicate_columns(_normalise_unvalidated_dataframe(df_unvalidated))
                    if "ObjectID" not in df_unval_save.columns and getattr(df_unval_save.index, "name", None) == "ObjectID":
                        df_unval_save = df_unval_save.reset_index()
                    if "Unvalidated_source" in existing_tables or "unvalidated_source" in existing_tables:
                        tname = "Unvalidated_source" if "Unvalidated_source" in existing_tables else "unvalidated_source"
                        cursor = conn.cursor()
                        cursor.execute(f'DELETE FROM "{tname}";')
                        df_unval_save.to_sql(tname, conn, if_exists="append", index=False)
                    else:
                        df_unval_save.to_sql("Unvalidated_source", conn, if_exists="replace", index=False)
                elif "Unvalidated_source" in existing_tables or "unvalidated_source" in existing_tables:
                    tname = "Unvalidated_source" if "Unvalidated_source" in existing_tables else "unvalidated_source"
                    cursor = conn.cursor()
                    cursor.execute(f'DELETE FROM "{tname}";')
        finally:
            conn.close()

    @staticmethod
    def import_from_excel(excel_path, sqlite_path, config):
        """Import data from an Excel file into a SQLite database.

        Creates a timestamped backup of the source Excel file in a 'backups'
        subdirectory before reading it, then writes the data into the SQLite
        database at sqlite_path.

        Args:
            excel_path:  Absolute path to the source .xlsx file.
            sqlite_path: Absolute path to the destination .db file.
            config:      Database config dict.

        Returns:
            (df_reg, df_obs, df_photo, df_log, df_unvalidated) — the data that was imported.
        """
        import config as _app_cfg
        prefs = _app_cfg.load_prefs() or {}
        advanced_prefs = prefs.get("advanced", {})
        enable_backup = prefs.get("enable_excel_import_backup", advanced_prefs.get("enable_excel_import_backup", True))
        
        if enable_backup:
            backup_dir = os.path.join(os.path.dirname(excel_path), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(
                backup_dir,
                os.path.basename(excel_path) + f".backup_{timestamp}.xlsx"
            )
            shutil.copy2(excel_path, backup_path)
            print(f"Backed up {excel_path} to {backup_path}")

        df_reg, df_obs, df_photo, df_log, df_unvalidated = ExcelRepository.load_excel(excel_path, config)
        SQLiteRepository.save_sqlite(sqlite_path, df_reg, df_obs, df_photo, df_log, df_unvalidated)
        return df_reg, df_obs, df_photo, df_log, df_unvalidated

    @staticmethod
    def generate_empty_dataframes(config):
        """Build empty dataframes with the correct columns for a new database.

        The column structure is derived from the config's ui_sections so that a
        freshly created file will accept data without needing schema migrations.

        Args:
            config: Database config dict.

        Returns:
            (df_reg, df_obs, df_log, df_unvalidated) — four empty DataFrames with correct columns.
        """
        sections = config.get("ui_sections", {})

        reg_columns = ["ObjectID"]
        seen_reg = {"ObjectID"}
        for field in sections.get("registration", []):
            if field["name"] not in seen_reg:
                seen_reg.add(field["name"])
                reg_columns.append(field["name"])

        obs_columns = ["ObjectID"]
        seen_obs = {"ObjectID"}
        for field in sections.get("location", []):
            if field["name"] not in seen_obs:
                seen_obs.add(field["name"])
                obs_columns.append(field["name"])
        for field in sections.get("problems", []):
            if field["name"] not in seen_obs:
                seen_obs.add(field["name"])
                obs_columns.append(field["name"])

        obs_columns.extend([
            "Images_Missing", "Images_Problem", "Images_Wrong",
            ONLINE_EXISTS_COLUMN, REVIEWED_COLUMN, REVIEWED_AT_COLUMN
        ])

        df_reg = pd.DataFrame(columns=reg_columns)
        df_obs = pd.DataFrame(columns=obs_columns)
        df_log = _normalise_log_dataframe(pd.DataFrame())
        df_unvalidated = _normalise_unvalidated_dataframe(pd.DataFrame())
        return df_reg, df_obs, df_log, df_unvalidated

    @staticmethod
    def export_to_excel(sqlite_path, excel_path, config, progress_callback=None,
                        df_reg=None, df_obs=None, df_log=None, df_photo=None, df_unvalidated=None):
        """Export data to an Excel (.xlsx) file.

        If df_reg / df_obs are not provided, the data is read from sqlite_path.
        If sqlite_path also does not exist, empty dataframes are used (allowing
        creation of a blank template file).

        The Registration, Observation, Photo (if present), Log, and Unvalidated_source (if present) sheets are written.

        Args:
            sqlite_path:       Path to the source .db file (may be None).
            excel_path:        Path to write the output .xlsx file.
            config:            Database config dict.
            progress_callback: Optional callable(current, total, label) for
                                progress reporting.
            df_reg:            Pre-loaded Registration dataframe (optional).
            df_obs:            Pre-loaded Observation dataframe (optional).
            df_log:            Pre-loaded Log dataframe (optional).
            df_photo:          Pre-loaded Photo dataframe (optional).
            df_unvalidated:    Pre-loaded Unvalidated source dataframe (optional).
        """
        if df_reg is None or df_obs is None:
            if sqlite_path and os.path.exists(sqlite_path):
                df_reg, df_obs, df_photo_loaded, df_log, df_unvalidated_loaded = SQLiteRepository.load_sqlite(sqlite_path, config)
                if df_photo is None:
                    df_photo = df_photo_loaded
                if df_unvalidated is None:
                    df_unvalidated = df_unvalidated_loaded
            else:
                df_reg, df_obs, df_log, df_unvalidated = SQLiteRepository.generate_empty_dataframes(config)

        if df_log is None or df_log.empty:
            df_log = _normalise_log_dataframe(pd.DataFrame())
        else:
            df_log = _normalise_log_dataframe(df_log)

        if df_unvalidated is None or df_unvalidated.empty:
            df_unvalidated = _normalise_unvalidated_dataframe(pd.DataFrame())
        else:
            df_unvalidated = _normalise_unvalidated_dataframe(df_unvalidated)

        from utils import sanitize_df_for_excel

        if df_reg is not None and "ObjectID" not in df_reg.columns:
            df_reg = df_reg.reset_index()
        
        if df_obs is not None and "ObjectID" not in df_obs.columns:
            df_obs = df_obs.reset_index()

        if df_photo is not None and not df_photo.empty and "ObjectID" not in df_photo.columns:
            df_photo = df_photo.reset_index()

        if df_unvalidated is not None and not df_unvalidated.empty and "ObjectID" not in df_unvalidated.columns:
            df_unvalidated = df_unvalidated.reset_index()

        # Sanitize dataframes to prevent CSV/Excel formula injection
        df_reg_safe = sanitize_df_for_excel(df_reg)
        df_obs_safe = sanitize_df_for_excel(df_obs)
        df_photo_safe = sanitize_df_for_excel(df_photo)
        df_log_safe = sanitize_df_for_excel(df_log)
        df_unval_safe = sanitize_df_for_excel(df_unvalidated) if df_unvalidated is not None and not df_unvalidated.empty else None

        sheets = config.get("sheets", {
            "reg": "Registration", "obs": "Observation",
            "photo": "Photo", "log": "Log", "unvalidated": "Unvalidated_source"
        })

        steps = [
            (df_reg_safe, sheets.get("reg", "Registration"), "Registration"),
            (df_obs_safe, sheets.get("obs", "Observation"),  "Observation"),
        ]
        if df_photo_safe is not None and not df_photo_safe.empty:
            steps.append((df_photo_safe, sheets.get("photo", "Photo"), "Photo"))
        steps.append((df_log_safe, sheets.get("log", "Log"), "Log"))
        if df_unval_safe is not None and not df_unval_safe.empty:
            steps.append((df_unval_safe, sheets.get("unvalidated", "Unvalidated_source"), "Unvalidated_source"))

        total = len(steps)

        # Always write mode via a temp file → atomic rename (never leaves
        # the destination in a partially-written / corrupt state, and avoids
        # openpyxl append-mode which corrupts ZIP structure on overwrite).
        parent_dir = os.path.dirname(os.path.abspath(excel_path))
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)
        temp_id = uuid.uuid4().hex[:8]
        tmp_path = f"{excel_path}.{temp_id}.tmp.xlsx"
        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl", mode="w") as writer:
                for i, (df, sheet, label) in enumerate(steps, 1):
                    if progress_callback:
                        progress_callback(i - 1, total, f"Writing {label}...")
                    if df is not None:
                        df.to_excel(writer, sheet_name=sheet, index=False)
                    if progress_callback:
                        progress_callback(i, total, f"{label} done")

                # Apply worksheet styling & autosizing
                try:
                    from openpyxl.utils import get_column_letter
                    for sheet_name, ws in writer.sheets.items():
                        # Freeze header row
                        ws.freeze_panes = "A2"
                        # Auto-filter
                        if ws.max_row > 1 and ws.max_column > 0:
                            ws.auto_filter.ref = ws.dimensions
                        # Auto-size columns (sample up to 500 rows for speed without loading entire columns)
                        max_sample_rows = min(ws.max_row, 500)
                        if max_sample_rows > 0 and ws.max_column > 0:
                            col_max_lens = {c: 0 for c in range(1, ws.max_column + 1)}
                            for row_cells in ws.iter_rows(min_row=1, max_row=max_sample_rows):
                                for col_idx, cell in enumerate(row_cells, 1):
                                    val_str = str(cell.value or "")
                                    if "\n" in val_str:
                                        val_str = max(val_str.split("\n"), key=len)
                                    if len(val_str) > col_max_lens[col_idx]:
                                        col_max_lens[col_idx] = len(val_str)
                            for col_idx, max_len in col_max_lens.items():
                                col_letter = get_column_letter(col_idx)
                                ws.column_dimensions[col_letter].width = max(10, min(max_len + 3, 60))
                except Exception as style_err:
                    debug_error("export_to_excel styling", str(style_err))

            # Atomic rename with retry loop for transient locks (OneDrive, Dropbox, antivirus)
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    os.replace(tmp_path, excel_path)
                    break
                except (PermissionError, OSError) as e:
                    if attempt == max_retries - 1:
                        try:
                            if os.path.exists(tmp_path):
                                os.remove(tmp_path)
                        except OSError:
                            pass
                        raise PermissionError(
                            f"Cannot save to '{excel_path}'. The file is open in Microsoft Excel or locked by OneDrive/Dropbox. Please close it or wait for sync to complete and retry."
                        ) from e
                    time.sleep(0.15 * (2 ** attempt))
        except PermissionError:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass
            raise
        except Exception:
            try:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)
            except OSError:
                pass
            raise
